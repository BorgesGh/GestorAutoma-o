import pyautogui as pg
import time
import openpyxl
import os

#Acessar área de inserção dos tempos

time.sleep(5)#Tempo para acessar outra tela, no caso o navegador

#A planilha de processos e tempos devem estar Na seguinte formatação (00:30) -> 30 minutos
# Coluna A  / Coluna B 
# Pintura  /   00:30 
# Trocar  /    01:30 
# ....        ......

# Obter o diretório atual do arquivo Python
diretorio_atual = os.path.dirname(os.path.abspath(__file__))

# Construir o caminho completo para a planilha
caminho_planilha = os.path.join(diretorio_atual, "Teste.xlsx")

# Carregar a planilha usando o caminho correto
planilhaProcessos = openpyxl.load_workbook(caminho_planilha)

sheet = planilhaProcessos['Sheet1']

OrdemDeServico = [] #Nome da OS e tempo
# =============================================================== ///////////////// =================================
def leituraDeProcessos(sheet):
    Processos = []

    linhas = sheet.max_row

    for i in range(1,linhas + 1):
        celula = sheet.cell(row = i,column = 1)
        Processos.append(celula.value)
        
    return Processos
# =============================================================== ///////////////// =================================
def leituraDeTempos(sheet):
    
    Tempos = []

    linhas = sheet.max_row

    for i in range(1,linhas + 1):
        celula = sheet.cell(row = i,column = 2)
        Tempos.append(celula.value + ":00")
        
    return Tempos
# =============================================================== ///////////////// =================================
def limparConteudo():
    pg.keyDown('ctrl')
    pg.press('a')
    pg.press('delete')
    pg.keyUp('ctrl')
    time.sleep(1)
# =============================================================== ///////////////// =================================
def adicionarProcesso(Processos):

    #Botao "novo"
    botaoNovo = pg.position(x=411, y=224)
    

    for i,e in enumerate(Processos):
        #Clicar no botão "Novo"
        
        pg.click(botaoNovo)
        #Digitar nome do processo...
        pg.write(Processos[i])
        #Pressionar Enter para salvar...
        pg.press('enter')
        #Agurdar alguns segundos de save...
        time.sleep(3)
# =============================================================== ///////////////// =================================      
def adicionarTempoProcesso(Processos, Tempos):


    #botao de selecionar o processo
    botaoProcesso = pg.position(x=752, y=291)

    #Barra de pesquisa
    barraPesquisa = pg.position(x=771, y=337)

    #Primeira opção dos processos
    primeiraOp = pg.position(x=782, y=396)

    #Area de tempo padrao
    areaTempo = pg.position(x=1715, y=343)

    #Botão salvar
    salvar = pg.position(x=1145, y=954)
    #Pressionando tab e seguida, é possivel cofirmar o tempo padrão
    
    for i,e in enumerate(Processos):

        pg.click(botaoProcesso)
        time.sleep(0.5)
        pg.write(Processos[i]) #Escreve o nome do processo como esta na planilha
        time.sleep(0.5)
        pg.click(primeiraOp)

        pg.click(areaTempo)
        pg.write(Tempos[i]) #Escreve o tempo padrão
        time.sleep(0.5)
        pg.press("tab") #Acessar o botão de confirmação
        pg.press("enter") 

        time.sleep(0.5)
        pg.click(salvar)
        time.sleep(3)

        #Desmarcar o processo
        pg.click(botaoProcesso)
        pg.click(primeiraOp)

        #Apagar o processo que havia escrito
        pg.click(barraPesquisa)
        limparConteudo()

        pg.click(botaoProcesso)
# =============================================================== ///////////////// =================================
def manipularSheet(Processos):
    
    for i,e in enumerate(Processos):
        Processos[i] = Processos[i].capitalize() #Transforma os processos com nome padrao
        sheet["A" + str(i + 1)] = Processos[i]

    planilhaProcessos.save('Teste.xlsx')

# =============================================================== ///////////////// =================================
def adicionarFeriados():

    #Menu retraido, 90% de zoom

    #Novo botao
    novo = pg.position(x=146, y=247)

    #Descricao
    descricao = pg.position(x=926, y=334)

    #Data (13/02)
    dataAba = pg.position(x=927, y=519)
    dataDia = pg.position(x=881, y=725)

    #Horario de inicio
    horaInicial = pg.position(x=813, y=616)

    #Duração
    duracao = pg.position(x=989, y=610)

    #Falta botão
    faltaBotao = pg.position(x=1302, y=614)

    pg.click(novo)
    time.sleep(2)
    pg.click(descricao)
    pg.write("Carnaval")

    pg.click(dataAba)
    pg.click(dataDia)
    
    time.sleep(2)
    pg.click(horaInicial)
    limparConteudo()
    pg.write("07:00")

    pg.click(duracao)
    time.sleep(1)
    pg.write("34:30")

    pg.click(faltaBotao)

    #Tempo para selecionar e salvar
    time.sleep(10)

    return 0

# =============================================================== ///////////////// =================================
def criarFormulario(Processos,questionario):
    #90% de zoom
    #Menu fechado

    #Novo
    buttonNovo = pg.position(x=68, y=263)

    #Aba de processos   
    processo = pg.position(x=273, y=325)

    #barra de pesquisa
    barraDePesquisa = pg.position(x=255, y=381)

    #primeira opção
    primeiraOpcao = pg.position(x=228, y=456)

    #Aba de tipo
    abaTipo = pg.position(x=1524, y=324)

    #Aba de descrição
    descricao = pg.position(x=319, y=429)

    #aba True ou false
    trueFalse = pg.position(x=675, y=444)


    #for i in Processos:



# =============================================================== //// -- MAIN --///// =================================
Processos = leituraDeProcessos(sheet)
Tempos = leituraDeTempos(sheet)

#Descomentar apenas o que será feito

#adicionarProcesso(Processos)
#adicionarTempoProcesso(Processos,Tempos)

questionario = [["Tipo","Descrição",True],["Tipo2","Descrição2",False]]

print(pg.position())
